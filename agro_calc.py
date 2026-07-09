import pandas as pd
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


AIR_COL = "avg_temp"

PHASES = [
    "Kurtakning bo'rtishi",
    "1-barg yozilishi",
    "Gullashi",
    "Mevaning shakllanishi",
    "Pishib yetilish boshlanishi"
]


def make_date(y, m, d):
    try:
        if pd.isna(y) or pd.isna(m) or pd.isna(d):
            return None
        return datetime(int(y), int(m), int(d))
    except Exception:
        return None


def date_text(dt):
    if dt is None or pd.isna(dt):
        return ""
    return f"{dt.day}/{dt.month}"


def read_agro_data(data_excel_path, station_name):
    xls = pd.ExcelFile(data_excel_path)

    sheet_map = {
        str(sheet).strip().lower(): sheet
        for sheet in xls.sheet_names
    }

    key = str(station_name).strip().lower()

    if key not in sheet_map:
        raise ValueError(
            f"'{station_name}' nomli stansiya bazada topilmadi! "
            f"Mavjud sheetlar: {', '.join(xls.sheet_names)}"
        )

    df = pd.read_excel(data_excel_path, sheet_name=sheet_map[key])
    df.columns = [str(c).strip().lower() for c in df.columns]

    required_cols = ["year", "month", "days", AIR_COL]

    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Harorat bazasida '{col}' ustuni topilmadi!")

    for col in required_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=required_cols)

    df["year"] = df["year"].astype(int)
    df["month"] = df["month"].astype(int)
    df["days"] = df["days"].astype(int)

    df["date"] = pd.to_datetime(
        {
            "year": df["year"],
            "month": df["month"],
            "day": df["days"]
        },
        errors="coerce"
    )

    df = df.dropna(subset=["date"])
    df = df.sort_values("date").reset_index(drop=True)

    return df


def read_phase_excel(phase_file):
    raw = pd.read_excel(phase_file, header=None)

    rows = []

    for i in range(1, len(raw)):
        r = raw.iloc[i]

        station = r.iloc[0]

        p1 = make_date(r.iloc[1], r.iloc[2], r.iloc[3])
        p2 = make_date(r.iloc[4], r.iloc[5], r.iloc[6])
        p3 = make_date(r.iloc[7], r.iloc[8], r.iloc[9])
        p4 = make_date(r.iloc[10], r.iloc[11], r.iloc[12])
        p5 = make_date(r.iloc[13], r.iloc[14], r.iloc[15])

        dates = [p1, p2, p3, p4, p5]
        years = [d.year for d in dates if d is not None]

        if not years:
            continue

        rows.append({
            "year": years[0],
            "station": station,
            "Kurtakning bo'rtishi": p1,
            "1-barg yozilishi": p2,
            "Gullashi": p3,
            "Mevaning shakllanishi": p4,
            "Pishib yetilish boshlanishi": p5
        })

    if not rows:
        raise ValueError("Fenologik Excel fayldan faza sanalari o'qilmadi!")

    return pd.DataFrame(rows)


def find_start_by_base_temp(df, year, base_temp):
    data = df[
        (df["year"] == year) &
        (df[AIR_COL] >= base_temp)
    ].copy()

    if data.empty:
        return None

    data = data.sort_values("date")

    return data.iloc[0]["date"]


def calc_period(df, start_date, end_date, base_temp):
    if start_date is None or end_date is None:
        return {
            "kun": 0,
            "aktiv": 0,
            "effektiv": 0
        }

    if end_date < start_date:
        return {
            "kun": 0,
            "aktiv": 0,
            "effektiv": 0
        }

    data = df[
        (df["date"] >= start_date) &
        (df["date"] <= end_date)
    ].copy()

    data = data.dropna(subset=[AIR_COL])

    selected = data[data[AIR_COL] >= base_temp]

    aktiv = selected[AIR_COL].sum()
    effektiv = (selected[AIR_COL] - base_temp).sum()

    return {
        "kun": int(len(selected)),
        "aktiv": round(float(aktiv), 1),
        "effektiv": round(float(effektiv), 1)
    }


def build_natija_sheet(df, phase_df, station_name, base_temp):
    rows = []

    for _, r in phase_df.iterrows():
        year = int(r["year"])
        start_date = find_start_by_base_temp(df, year, base_temp)

        station = r.get("station", station_name)

        if pd.isna(station) or str(station).strip() == "":
            station = station_name

        row = {
            "YILLAR": year,
            "Stansiya": station,
            f"Hisoblash boshlangan sana >= {base_temp}°C": date_text(start_date)
        }

        previous_date = start_date

        for phase in PHASES:
            phase_date = r[phase]

            interval_calc = calc_period(
                df=df,
                start_date=previous_date,
                end_date=phase_date,
                base_temp=base_temp
            )

            season_calc = calc_period(
                df=df,
                start_date=start_date,
                end_date=phase_date,
                base_temp=base_temp
            )

            row[f"{phase} | Sana"] = date_text(phase_date)
            row[f"{phase} | Kun - Havo"] = interval_calc["kun"]
            row[f"{phase} | Havo Aktiv ΣT"] = interval_calc["aktiv"]
            row[f"{phase} | Havo Effektiv ΣT"] = interval_calc["effektiv"]
            row[f"{phase} | Mavsum boshidan Aktiv ΣT"] = season_calc["aktiv"]
            row[f"{phase} | Mavsum boshidan Effektiv ΣT"] = season_calc["effektiv"]

            previous_date = phase_date

        rows.append(row)

    return pd.DataFrame(rows)


def build_fazalar_farqi_sheet(df, phase_df, station_name, base_temp):
    rows = []

    for _, r in phase_df.iterrows():
        year = int(r["year"])
        start_date = find_start_by_base_temp(df, year, base_temp)

        station = r.get("station", station_name)

        if pd.isna(station) or str(station).strip() == "":
            station = station_name

        dates = {
            f"Boshlanish >= {base_temp}°C": start_date,
            "Kurtakning bo'rtishi": r["Kurtakning bo'rtishi"],
            "1-barg yozilishi": r["1-barg yozilishi"],
            "Gullashi": r["Gullashi"],
            "Mevaning shakllanishi": r["Mevaning shakllanishi"],
            "Pishib yetilish boshlanishi": r["Pishib yetilish boshlanishi"]
        }

        pairs = [
            (f"Boshlanish >= {base_temp}°C", "Kurtakning bo'rtishi"),
            ("Kurtakning bo'rtishi", "1-barg yozilishi"),
            ("1-barg yozilishi", "Gullashi"),
            ("Gullashi", "Mevaning shakllanishi"),
            ("Mevaning shakllanishi", "Pishib yetilish boshlanishi")
        ]

        row = {
            "YILLAR": year,
            "Stansiya": station
        }

        for start_name, end_name in pairs:
            s_date = dates[start_name]
            e_date = dates[end_name]

            calc = calc_period(
                df=df,
                start_date=s_date,
                end_date=e_date,
                base_temp=base_temp
            )

            block = f"{start_name} -> {end_name}"

            row[f"{block} | Boshlanish sana"] = date_text(s_date)
            row[f"{block} | Tugash sana"] = date_text(e_date)
            row[f"{block} | Kun - Havo"] = calc["kun"]
            row[f"{block} | Havo Aktiv ΣT"] = calc["aktiv"]
            row[f"{block} | Havo Effektiv ΣT"] = calc["effektiv"]

        rows.append(row)

    return pd.DataFrame(rows)


def format_excel(output_path):
    wb = load_workbook(output_path)

    green = PatternFill("solid", fgColor="C6E0B4")
    blue = PatternFill("solid", fgColor="D9EAF7")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="999999")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            value = str(cell.value)

            if "sana" in value.lower():
                cell.fill = blue
            elif "Mavsum boshidan" in value:
                cell.fill = yellow
            else:
                cell.fill = green

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    wb.save(output_path)


def create_result_excel(
    data_excel_path,
    phase_file,
    station_name,
    output_path,
    base_temp=10
):
    base_temp = int(base_temp)

    df = read_agro_data(
        data_excel_path=data_excel_path,
        station_name=station_name
    )

    phase_df = read_phase_excel(phase_file)

    natija_df = build_natija_sheet(
        df=df,
        phase_df=phase_df,
        station_name=station_name,
        base_temp=base_temp
    )

    farq_df = build_fazalar_farqi_sheet(
        df=df,
        phase_df=phase_df,
        station_name=station_name,
        base_temp=base_temp
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        natija_df.to_excel(
            writer,
            sheet_name="Natija",
            index=False
        )

        farq_df.to_excel(
            writer,
            sheet_name="Fazalar_farqi",
            index=False
        )

    format_excel(output_path)

    return output_path